"""IO 扩展: JSON / Excel / Parquet / Pickle / SQL 读写。

所有函数都接受/返回 DataFrame，与 pandas IO API 对齐。
"""

from __future__ import annotations

import json as _json
import pickle as _pickle
from typing import Any, Dict, Optional, Union
from .dataframe import DataFrame
# from .series import Series


# ============================================================================
# JSON
# ============================================================================

def read_json(
    path: str,
    orient: str = "records",
    lines: bool = False,
    encoding: str = "utf-8",
) -> DataFrame:
    """从 JSON 文件读取 DataFrame。

    Parameters
    ----------
    path : str
        JSON 文件路径。
    orient : str, default 'records'
        JSON 格式方向：
        - 'records': list[dict] (每行一个 dict)
        - 'columns': dict[str, list] (每列一个 list)
        - 'index': dict[str, dict] (行索引 → 列值)
        - 'split': {'columns': [...], 'data': [[...], ...]}
        - 'values': list[list] (纯二维数组)
    lines : bool, default False
        是否按行读取 JSON (每行一个 JSON 对象)。
    encoding : str, default 'utf-8'
        文件编码。

    Returns
    -------
    DataFrame
    """
    with open(path, "r", encoding=encoding) as f:
        if lines:
            records = [_json.loads(line) for line in f if line.strip()]
            return DataFrame(records)
        raw = _json.load(f)

    if orient == "records":
        return DataFrame(raw)
    elif orient == "columns":
        return DataFrame(raw)
    elif orient == "index":
        records = []
        for idx, row_dict in raw.items():
            record = {"index": idx, **row_dict}
            records.append(record)
        return DataFrame(records)
    elif orient == "split":
        cols = raw.get("columns", [])
        data = raw.get("data", [])
        return DataFrame(data, columns=cols)
    elif orient == "values":
        return DataFrame(raw)
    else:
        raise ValueError(f"Unknown orient: {orient}")


def to_json(
    df: DataFrame,
    path: Optional[str] = None,
    orient: str = "records",
    lines: bool = False,
    force_ascii: bool = False,
    indent: Optional[int] = None,
) -> Optional[str]:
    """将 DataFrame 写入 JSON 文件或返回 JSON 字符串。

    Parameters
    ----------
    df : DataFrame
        要写入的 DataFrame。
    path : str, optional
        输出文件路径。None 则返回字符串。
    orient : str, default 'records'
        JSON 格式方向。
    lines : bool, default False
        是否按行输出 JSON。
    force_ascii : bool, default False
        是否强制 ASCII 编码。
    indent : int, optional
        缩进空格数。

    Returns
    -------
    str or None
    """
    # df.values 返回 list[dict]
    records = df.values

    if orient == "records":
        data = records
    elif orient == "columns":
        data = {col: [row.get(col) for row in records] for col in df.columns}
    elif orient == "index":
        data = {}
        for i, row in enumerate(records):
            data[str(i)] = row
    elif orient == "split":
        data = {
            "columns": list(df.columns),
            "data": [[row.get(c) for c in df.columns] for row in records],
        }
    elif orient == "values":
        data = [[row.get(c) for c in df.columns] for row in records]
    else:
        raise ValueError(f"Unknown orient: {orient}")

    json_kwargs: Dict[str, Any] = {"ensure_ascii": force_ascii}
    if indent is not None:
        json_kwargs["indent"] = indent

    if lines:
        if orient != "records":
            raise ValueError("lines=True requires orient='records'")
        output = "\n".join(_json.dumps(r, **json_kwargs) for r in data)
    else:
        output = _json.dumps(data, **json_kwargs)

    if path is not None:
        with open(path, "w", encoding="utf-8") as f:
            f.write(output)
        if not output.endswith("\n"):
            with open(path, "a", encoding="utf-8") as f:
                f.write("\n")
        return None
    return output


# ============================================================================
# Excel
# ============================================================================

def read_excel(
    path: str,
    sheet_name: Union[str, int] = 0,
    header: int = 0,
    **kwargs,
) -> DataFrame:
    """从 Excel 文件读取 DataFrame。

    Parameters
    ----------
    path : str
        Excel 文件路径 (.xlsx / .xls)。
    sheet_name : str or int, default 0
        工作表名称或索引。
    header : int, default 0
        用作列名的行号。
    **kwargs
        传递给 openpyxl/pandas 的其他参数。

    Returns
    -------
    DataFrame
    """
    try:
        import openpyxl  # noqa: F401
        return _read_excel_openpyxl(path, sheet_name, header)
    except ImportError:
        pass

    try:
        import pandas as pd
        pdf = pd.read_excel(path, sheet_name=sheet_name, header=header, **kwargs)
        return DataFrame.from_pandas(pdf)
    except ImportError:
        raise ImportError(
            "read_excel requires openpyxl or pandas to be installed. "
            "Install with: pip install openpyxl"
        )


def _read_excel_openpyxl(
    path: str,
    sheet_name: Union[str, int] = 0,
    header: int = 0,
) -> DataFrame:
    """使用 openpyxl 读取 Excel。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return DataFrame()

    # 确定列名
    if header < len(rows):
        col_names = [
            str(c) if c is not None else f"col{i}"
            for i, c in enumerate(rows[header])
        ]
        data_rows = rows[header + 1:]
    else:
        col_names = [f"col{i}" for i in range(len(rows[0]))]
        data_rows = rows

    # 去重列名
    seen: Dict[str, int] = {}
    final_cols = []
    for c in col_names:
        if c in seen:
            seen[c] += 1
            final_cols.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            final_cols.append(c)

    # 构建数据
    data: Dict[str, list] = {c: [] for c in final_cols}
    for row in data_rows:
        for i, c in enumerate(final_cols):
            data[c].append(row[i] if i < len(row) else None)

    return DataFrame(data)


def to_excel(
    df: DataFrame,
    path: str,
    sheet_name: str = "Sheet1",
    index: bool = False,
    header: bool = True,
    **kwargs,
) -> None:
    """将 DataFrame 写入 Excel 文件。

    Parameters
    ----------
    df : DataFrame
        要写入的 DataFrame。
    path : str
        输出文件路径。
    sheet_name : str, default 'Sheet1'
        工作表名称。
    index : bool, default False
        是否写入行索引。
    header : bool, default True
        是否写入列名。
    **kwargs
        传递给 openpyxl/pandas 的其他参数。
    """
    try:
        import openpyxl  # noqa: F401
        _to_excel_openpyxl(df, path, sheet_name, index, header)
        return
    except ImportError:
        pass

    try:
        pdf = df.to_pandas()
        pdf.to_excel(path, sheet_name=sheet_name, index=index, header=header, **kwargs)
        return
    except ImportError:
        raise ImportError(
            "to_excel requires openpyxl or pandas to be installed. "
            "Install with: pip install openpyxl"
        )


def _to_excel_openpyxl(
    df: DataFrame,
    path: str,
    sheet_name: str = "Sheet1",
    index: bool = False,
    header: bool = True,
) -> None:
    """使用 openpyxl 写入 Excel。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写入表头
    if header:
        col_offset = 1 if index else 0
        for j, col_name in enumerate(df.columns):
            ws.cell(row=1, column=j + col_offset + 1, value=col_name)

    # 写入数据
    values = df.values
    row_start = 2 if header else 1
    for i, row in enumerate(values):
        for j, val in enumerate(row):
            col_offset = 1 if index else 0
            ws.cell(row=i + row_start, column=j + col_offset + 1, value=val)

    # 写入索引
    if index:
        for i in range(len(values)):
            ws.cell(row=i + row_start, column=1, value=i)

    wb.save(path)


# ============================================================================
# Parquet
# ============================================================================

def read_parquet(path: str, **kwargs) -> DataFrame:
    """从 Parquet 文件读取 DataFrame。

    Parameters
    ----------
    path : str
        Parquet 文件路径。
    **kwargs
        传递给 pyarrow/pandas 的其他参数。

    Returns
    -------
    DataFrame
    """
    try:
        import pyarrow.parquet as pq
        table = pq.read_table(path, **kwargs)
        return _arrow_table_to_dataframe(table)
    except ImportError:
        pass

    try:
        import pandas as pd
        pdf = pd.read_parquet(path, **kwargs)
        return DataFrame.from_pandas(pdf)
    except ImportError:
        raise ImportError(
            "read_parquet requires pyarrow or pandas to be installed. "
            "Install with: pip install pyarrow"
        )


def _arrow_table_to_dataframe(table) -> DataFrame:
    """将 PyArrow Table 转换为 DataFrame。"""
    data: Dict[str, list] = {}
    for col_name in table.column_names:
        col = table.column(col_name)
        data[col_name] = col.to_pylist()
    return DataFrame(data)


def to_parquet(
    df: DataFrame,
    path: str,
    compression: Optional[str] = "snappy",
    **kwargs,
) -> None:
    """将 DataFrame 写入 Parquet 文件。

    Parameters
    ----------
    df : DataFrame
        要写入的 DataFrame。
    path : str
        输出文件路径。
    compression : str, optional, default 'snappy'
        压缩算法 (snappy, gzip, brotli, zstd, none)。
    **kwargs
        传递给 pyarrow/pandas 的其他参数。
    """
    try:
        import pyarrow.parquet as pq

        table = _dataframe_to_arrow_table(df)
        pq.write_table(table, path, compression=compression, **kwargs)
        return
    except ImportError:
        pass

    try:
        pdf = df.to_pandas()
        pdf.to_parquet(path, compression=compression, **kwargs)
        return
    except ImportError:
        raise ImportError(
            "to_parquet requires pyarrow or pandas to be installed. "
            "Install with: pip install pyarrow"
        )


def _dataframe_to_arrow_table(df: DataFrame):
    """将 DataFrame 转换为 PyArrow Table。"""
    import pyarrow as pa

    arrays = []
    for col_name in df.columns:
        col_data = list(df[col_name].values)
        # 推断类型
        non_null = [v for v in col_data if v is not None]
        if not non_null:
            arrays.append(pa.array(col_data, type=pa.string()))
        elif all(isinstance(v, bool) for v in non_null):
            arrays.append(pa.array(col_data, type=pa.bool_()))
        elif all(isinstance(v, int) for v in non_null):
            arrays.append(pa.array(col_data, type=pa.int64()))
        elif all(isinstance(v, float) for v in non_null):
            arrays.append(pa.array(col_data, type=pa.float64()))
        else:
            arrays.append(pa.array([str(v) if v is not None else None for v in col_data]))

    return pa.table(dict(zip(df.columns, arrays)))


# ============================================================================
# Pickle
# ============================================================================

def read_pickle(path: str, **kwargs) -> DataFrame:
    """从 Pickle 文件读取 DataFrame。

    Parameters
    ----------
    path : str
        Pickle 文件路径。
    **kwargs
        传递给 pickle.load 的其他参数。

    Returns
    -------
    DataFrame
    """
    import pandas as pd

    try:
        pdf = pd.read_pickle(path, **kwargs)
        return DataFrame.from_pandas(pdf)
    except Exception:
        # 尝试直接用 pickle 加载
        with open(path, "rb") as f:
            obj = _pickle.load(f)
        if isinstance(obj, DataFrame):
            return obj
        raise TypeError(f"Pickle file contains {type(obj).__name__}, not DataFrame")


def to_pickle(df: DataFrame, path: str, **kwargs) -> None:
    """将 DataFrame 写入 Pickle 文件。

    Parameters
    ----------
    df : DataFrame
        要写入的 DataFrame。
    path : str
        输出文件路径。
    **kwargs
        传递给 pickle.dump 的其他参数。
    """
    try:
        pdf = df.to_pandas()
        pdf.to_pickle(path, **kwargs)
        return
    except ImportError:
        pass

    # 直接用 pickle 序列化 DataFrame
    with open(path, "wb") as f:
        _pickle.dump(df, f, **kwargs)


# ============================================================================
# SQL
# ============================================================================

def read_sql(
    query: str,
    conn,
    **kwargs,
) -> DataFrame:
    """从 SQL 数据库读取 DataFrame。

    Parameters
    ----------
    query : str
        SQL 查询语句。
    conn : sqlalchemy Engine 或 Connection
        数据库连接。
    **kwargs
        传递给 pandas.read_sql 的其他参数。

    Returns
    -------
    DataFrame
    """
    try:
        import sqlalchemy  # noqa: F401
    except ImportError:
        raise ImportError(
            "read_sql requires sqlalchemy to be installed. "
            "Install with: pip install sqlalchemy"
        )

    try:
        import pandas as pd
        pdf = pd.read_sql(query, conn, **kwargs)
        return DataFrame.from_pandas(pdf)
    except ImportError:
        raise ImportError(
            "read_sql requires pandas to be installed. "
            "Install with: pip install pandas"
        )


def to_sql(
    df: DataFrame,
    name: str,
    conn,
    if_exists: str = "fail",
    index: bool = False,
    **kwargs,
) -> None:
    """将 DataFrame 写入 SQL 数据库。

    Parameters
    ----------
    df : DataFrame
        要写入的 DataFrame。
    name : str
        目标表名。
    conn : sqlalchemy Engine 或 Connection
        数据库连接。
    if_exists : str, default 'fail'
        表已存在时的行为：'fail', 'replace', 'append'。
    index : bool, default False
        是否写入行索引。
    **kwargs
        传递给 pandas.to_sql 的其他参数。
    """
    try:
        import sqlalchemy  # noqa: F401
    except ImportError:
        raise ImportError(
            "to_sql requires sqlalchemy to be installed. "
            "Install with: pip install sqlalchemy"
        )

    try:
        pdf = df.to_pandas()
        pdf.to_sql(name, conn, if_exists=if_exists, index=index, **kwargs)
        return
    except ImportError:
        raise ImportError(
            "to_sql requires pandas to be installed. "
            "Install with: pip install pandas"
        )
